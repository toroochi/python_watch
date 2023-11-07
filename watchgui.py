import PySimpleGUI as sg
import datetime
import openpyxl

book = openpyxl.load_workbook('excel_data/Book1.xlsx')
ws = book["memory"]
i = 1
j = 1
p = 1
q = 1

a = False
b = False
c = False 
d = False

wtime = 0
btime = 0
normTime = 0
finish = False

while True:
     if ws.cell(row=i,column=1).value != None:
          i += 1
     else:
          a = True
     if ws.cell(row=j,column=2).value != None:
          j += 1
     else:
          b = True
     if ws.cell(row=p,column=3).value != None:
          p += 1
     else:
          c = True
     if ws.cell(row=q,column=4).value != None:
          q += 1
     else:
          d = True
     if a == True and b == True and c == True and d == True:
          break

sg.theme('DarkPurple1')
wstart = 0 #workstart
wend = 0
bstart = 0 #breakstart
bend = 0

wsbool = False
webool = False
bsbool = False
bebool = False
count = 0

def getTime():
     return datetime.datetime.now().strftime('%-m/%-d %H:%M:%S')

layout = [[sg.Text('', size=(20, 2), key='_time_',font=(40))],[sg.Text('Command'),sg.InputText(size=(20,3),key='-command-'),sg.Button('OK', key='-btn-')]]
window = sg.Window('Clock', layout, grab_anywhere=True)

def Record():
     global i,j,p,q,count,wsbool,webool,bsbool,bebool,finish
     if wsbool == True and count == 1:
          print(type(wstart))
          ws.cell(row=i,column=1).value = wstart
          print(wstart)
          i += 1
          wsbool = False
          book.save('excel_data/Book1.xlsx')
          print(wsbool)
          print("aaa")
          count = 0
     if webool == True and count == 1:
          ws.cell(row=j,column=2).value = wend
          j += 1
          webool = False
          book.save('excel_data/Book1.xlsx')
          count = 0
     if bsbool == True and count == 1:
          ws.cell(row=p,column=3).value = bstart
          p += 1
          bsbool = False
          book.save('excel_data/Book1.xlsx')
          count = 0
     if bebool == True and count == 1:
          ws.cell(row=q,column=4).value = bend
          q += 1
          bebool = False
          book.save('excel_data/Book1.xlsx')
          count = 0
     if finish == True:
          maxVal = [] 
          maxVal = [val for val in maxVal if val is not None]  # Noneを除外
          btime = abs((ws.cell(row=q - 1,column=4).value - ws.cell(row=p -1 ,column=3).value) - (ws.cell(row=1,column=4).value - ws.cell(row=1,column=3).value))
          if maxVal:  # maxValが空でないことを確認
               wtime = (max(maxVal) - min(maxVal)) - btime
               normTime = abs(wtime/btime - 52/17)/max(wtime/btime,52/17)
          else:
               wtime = 0
               normTime = 0

          print(wtime,btime,normTime)

          ws.cell(row=1,column=5).value = wtime
          ws.cell(row=2,column=5).value = btime
          ws.cell(row=3,column=5).value = normTime
          book.save('excel_data/Book1.xlsx')

          finish = False
     
while True:
     Record()
     event, values = window.Read(timeout=1000)
     if event == '-btn-':
          if values['-command-'] == 'wstart' and count == 0:
               wstart1 = datetime.datetime.now().time()
               wstart = (wstart1.hour * 100) + (wstart1.minute)
               window['-command-'].update('')
               wsbool = True
               count += 1
               print(count)
          elif values['-command-'] == 'wend' and count == 0:
               wend1 = datetime.datetime.now().time()
               wend = (wend1.hour * 100) + (wend1.minute)
               window['-command-'].update('')
               webool = True
               count += 1
          elif values['-command-'] == 'bstart' and count == 0:
               bstart1 = datetime.datetime.now().time()
               bstart = (bstart1.hour * 100) + (bstart1.minute)
               window['-command-'].update('')
               bsbool = True
               count += 1
          elif values['-command-'] == 'bend' and count == 0:
               bend1 = datetime.datetime.now().time()  
               bend = (bend1.hour * 100) + (bend1.minute)
               window['-command-'].update('')
               bebool = True
               count += 1
          elif values['-command-'] == 'finish':
               finish = True
               window['-command-'].update('')

     if event == sg.WINDOW_CLOSED:
         break
     window.find_element('_time_').Update(getTime())


window.Close()
